import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import os

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def main():
    root_window = tk.Tk()
    root_window.title("Excel IO")

    options = ("Excel Workbook to Text file (.xlsx to .txt)", "Text file(s) to Excel Workbook (.txt to .xlsx)")

    dropvar = tk.StringVar(root_window)
    dropvar.set(options[0])

    def openfiles(f):
        """Invoked from postprocess function to open the created files."""

        if isinstance(f, str):
            os.system("start " + f)

        else:
            for file in f:
                os.system("start " + "output/{0}".format(file))

    def callback():
        """Calls functions for converting files."""

        currentchoice = dropvar.get()

        if currentchoice == options[0]:
            accepted_file_types = [("Excel Workbooks", ".xlsx")]
            file = filedialog.askopenfilename(parent=root_window, initialdir=os.getcwd(), title="Select a file to be converted.", filetypes=accepted_file_types)

            if file is "":
                messagebox.showerror("Error", "No file was selected.")
            else:
                xlsx_to_txt(file)
                root_window.withdraw()
                postprocess(currentchoice)

        elif currentchoice == options[1]:
            accepted_file_types = [("Text Files", ".txt")]
            files = filedialog.askopenfilenames(parent=root_window, initialdir=os.getcwd(), title="Select a file to be converted. (Multiple can be chosen!)", filetypes=accepted_file_types)

            if len(files) == 0:
                messagebox.showerror("Error", "No file was selected.")
            else:
                txt_to_xlsx(files)
                root_window.withdraw()
                postprocess(currentchoice)

    def postprocess(choice):
        """Shows as a landing window for when the program is done converting files."""

        # I want to clear the previous window's widgets to make space for new ones.
        for widget in root_window.winfo_children():
            widget.destroy()

        root_window.deiconify()

        # TODO - Some of these windows look a bit awkward, need to fix that.

        mylabel = tk.Label(root_window, text="Conversion complete.")
        mylabel.grid(row=1, column=3, padx=150, pady=(25, 10))

        if choice == options[0]:
            filelist = []
            for file in os.listdir("output"):
                if file.endswith(".txt"):
                    filelist.append(file)

            mylabel2 = tk.Label(root_window, text="Files created:\n {0}".format("\n".join(filelist)))
            mylabel2.grid(row=2, column=3, padx=150, pady=(10, 10))

            mybutton = tk.Button(root_window, text="Click here to open the created files.", command=lambda: openfiles(filelist))
            mybutton.grid(row=3, column=3, padx=150, pady=(25, 65))

        elif choice == options[1]:
            mylabel2 = tk.Label(root_window, text="File created: {0}".format("workbook.xlsx"))
            mylabel2.grid(row=2, column=3, padx=150, pady=(10, 10))

            mybutton = tk.Button(root_window, text="Click here to open the created files.",  command=lambda: openfiles("output/workbook.xlsx"))
            mybutton.grid(row=3, column=3, padx=150, pady=(25, 65))

        center(root_window)

    mylabel = tk.Label(root_window, text="Choose a file conversion")
    mylabel.grid(row=1, column=3, padx=150, pady=(25, 10))

    optionprompt = tk.OptionMenu(root_window, dropvar, *options)
    optionprompt.grid(row=2, column=3, padx=150, pady=(0, 25))

    mybutton = tk.Button(root_window, text="Continue", command=callback)
    mybutton.grid(row=3, column=3, padx=150, pady=(25, 25))

    center(root_window)

    root_window.mainloop()


# Centers tk window. I had no idea how to do this. Code from https://bit.ly/2IumGmq
def center(win):
    win.update_idletasks()
    width = win.winfo_width()
    height = win.winfo_height()
    x = (win.winfo_screenwidth() // 2) - (width // 2)
    y = (win.winfo_screenheight() // 2) - (height // 2)
    win.geometry('{}x{}+{}+{}'.format(width, height, x, y))


def xlsx_to_txt(file):
    """Converts the given excel spreadsheet to text files. Splits into files for each column."""

    wb = openpyxl.load_workbook(file)

    workingsheet = wb.active

    maxcolumn_letter = get_column_letter(workingsheet.max_column)

    maxcolumn_index = column_index_from_string(maxcolumn_letter)

    # Loops over every column with content and writes each individual column to a file.
    # TODO - Make this work for rows too, but not sure how I'd detect that. Maybe ask the user?
    for i in range(1, maxcolumn_index + 1):
        with open("output/column{0}.txt".format(i), "w") as f:
            for cell in workingsheet[get_column_letter(i)]:
                f.write(str(cell.value))
                f.write("\n")


def txt_to_xlsx(files):
    """Converts the given text files to an excel spreadsheet. Can take several text files at once."""

    try:
        wb = openpyxl.load_workbook("output/workbook.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    workingsheet = wb.active

    workingcolumn = 1

    for file in files:
        with open(file, "r") as f:
            lines = f.readlines()

        # Checks for empty column
        if workingsheet["{0}{1}".format(get_column_letter(workingsheet.max_column), workingsheet.max_row)].value is not None:
            workingcolumn += 1

        counter = 1

        for x in lines:
            workingsheet["{0}{1}".format(get_column_letter(workingcolumn), str(counter))] = x
            counter += 1

    # Resize columns automatically
    # Had no idea how to do this. Code is from https://bit.ly/2tHaZDg , with a few of my personal edits.
    for col in workingsheet.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception as e:
                print(e)
                pass
        adjusted_width = (max_length + 2) * 1.2
        workingsheet.column_dimensions[column].width = adjusted_width

    try:
        wb.save("output/workbook.xlsx")

    except PermissionError:
        messagebox.showerror("Error", "The file is currently open.")

    except Exception as e:
        messagebox.showerror("Error", "An unexpected error has occurred: {0}".format(e))
        print(e)


if __name__ == "__main__":
    main()

